from openpyxl import load_workbook
import os

from conftest import RESOURCES_DIR, XLSX_NAME


def test_xlsx():
    xlsx_path = os.path.join(RESOURCES_DIR, XLSX_NAME)
    book = load_workbook(xlsx_path)
    sheet = book.active
    test_cell = sheet.cell(row=3, column=2)
    print(test_cell.value)

    assert test_cell.value == "Mara"
